import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from openpyxl.styles import PatternFill, colors
from progressbar import progressbar

red_background = PatternFill("solid", start_color=colors.RED, end_color=colors.RED)
green_background = PatternFill("solid", start_color=colors.GREEN, end_color=colors.GREEN)


class Reader:

    def __init__(self, inventory_filename: str, replenishment_filename: str):
        self.replenishment_filename = replenishment_filename
        self.inventory_book = load_workbook(inventory_filename, read_only=True)
        self.replenishment_book = load_workbook(replenishment_filename)
        self.inventory_sheet = self.inventory_book.active
        self.replenishment_sheet = self.replenishment_book.active
        self.quantity_on_hand_column_number, self.inventory_sku_columns = self.find_inventory_columns()
        self.replenishment_sku_column_number = self.find_column_number(self.replenishment_sheet, 'sku')

    def find_inventory_columns(self):
        sku_columns = []
        headers = next(self.inventory_sheet.iter_rows(min_row=1, max_row=1))
        quantity_on_hand_column_number = None
        for column in headers:
            value = column.value
            if value is None or not column.data_type == 's':
                continue
            else:
                if 'sku' in value.lower():
                    sku_columns.append(column.column - 1)
                if 'quantity available' == value.lower():
                    quantity_on_hand_column_number = column.column - 1
        if quantity_on_hand_column_number is None:
            raise Exception(f"Unable to find SKU Column in {self.replenishment_filename} ")
        return quantity_on_hand_column_number, set(sku_columns)

    def find_column_number(self,sheet, column_name):
        for head in sheet.iter_rows(1, 1):
            for column in head:
                if column_name in column.value.lower() and column.value.lower() != "fnsku":
                    return column.column
        else:
            raise Exception(f"Unable to find {column_name} in {self.replenishment_filename}")

    def run(self):
        print("building Inventory...")
        inventory = {}
        columns_itr = next(self.replenishment_sheet.iter_cols(self.replenishment_sku_column_number, min_row=2))
        data = {cell.value: cell.row for cell in columns_itr}
        for row in self.inventory_sheet.iter_rows(min_row=2):
            if isinstance(row, EmptyCell):
                continue
            columns = [row[x] for x in self.inventory_sku_columns if not isinstance(row[x], EmptyCell)]
            inventory.update({column.value: {'quantity': row[self.quantity_on_hand_column_number].value} for column in
                              columns if column.value in data})

        print("filling in colors...")

        for sku, item_data in progressbar(inventory.items()):
            row_index = data[sku]
            quantity = int(item_data['quantity'])
            row = self.replenishment_sheet[f"{row_index}:{row_index}"]
            for column in row:
                column.fill = green_background if quantity > 0 else red_background

        today = datetime.now().strftime('%a-%b-%d')
        self.replenishment_book.save(f'updated_replenishment - {today}.xlsx')


if __name__ == '__main__':
    try:
        inventory_files = os.listdir('inventory')
        replenishment_files = os.listdir('replenishment')
        reader = Reader(os.path.join('inventory', inventory_files[0]),
                        os.path.join('replenishment', replenishment_files[0]))
        reader.run()
    except Exception as e:
        print(f'Error has occurred: {e}')
        input('Press Enter to exit')
